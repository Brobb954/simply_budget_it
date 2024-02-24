import re
from openpyxl import Workbook, load_workbook
from contextlib import contextmanager
import os

@contextmanager
def managed_workbook(filename):
    if os.path.exists(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()
    try:
        yield workbook
    finally:
        workbook.save(filename)

class DataManager:
    def __init__(self, filename):
        self.filename = filename
        self.in_memory_data = [] # In-memory storage for data

    def load_or_create_workbook(self):
        try:
            with managed_workbook(self.filename) as workbook:
                self.workbook = workbook
                self.sheet = workbook.active
                if os.path.exists(self.filename):
                    self.load_data()
                    print("workbook loaded")
                else:
                    self.sheet["A1"] = "Type"
                    self.sheet["B1"] = "Description"
                    self.sheet["C1"] = "Amount"
                    print("New workbook created")
        except PermissionError:
            print("Permission denied: Unable to access the file")
        except Exception as e:
            print(f"An error occurred: {e}")

    def load_data(self):
        self.in_memory_data = []
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                self.in_memory_data.append(row)


    def save_to_excel(self):

        # Clear existing data if data exists
        if self.sheet.max_row > 1:
            self.sheet.delete_rows(2, self.sheet.max_row - 1)

        # Write new data from memory
        for entry in self.in_memory_data:
            self.sheet.append(entry)

        # Save workbook
        self.workbook.save(self.filename)
        print("saved")
        
    

    def delete_data(self, index):
        try:
            if index < len(self.in_memory_data):
                del self.in_memory_data[index]
            else:
                return "Error: No entry exists at the specified index"
            
        except Exception as e:
            return str(e)

    def clear_data(self):
        if len(self.in_memory_data) > 0:
           self.in_memory_data.clear()
           return "All entries have been cleared"
        else:
           return "There are no entries to delete"


    def update_data_entry(self, index, new_values):

        if 0 <= index < len(self.in_memory_data):
            self.in_memory_data[index] = new_values
            return "Entry updated successfully"
        else:
            return None
        
    def get_selected_entry(self, index):
        if 0 <= index < len(self.in_memory_data):
            return self.in_memory_data[index]
        else:
            return None
        

    def calculate_totals(self):
        # Initialize Totals
        total_income = 0.00
        total_expense = 0.00

        # Calculate totals from treeview data
        for entry in self.in_memory_data:
            if entry[0] == "Income":
                total_income+= float(entry[2].replace('$','').replace(',',''))
            else:
                total_expense += float(entry[2].replace('$','').replace(',',''))

        balance = total_income - total_expense
        return total_expense, total_income, balance
    
    def is_valid_amount(self, amount_str):
        pattern = r'\d+(\.\d{1,2})?$'
        return re.fullmatch(pattern, amount_str) is not None
    
    def format_entry(self, amount_str):
        if self.is_valid_amount(amount_str):
            amount_float = float(amount_str)
            formatted_amount = f"${amount_float:,.2f}"
            return(formatted_amount)
        else:
            return None
        
    def close(self):
        self.save_to_excel()


        